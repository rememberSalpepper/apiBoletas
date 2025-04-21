# main.py
import os
import io
import json
import re
import time  # Necesario para Gemini
from datetime import datetime
from typing import List

import pandas as pd
from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from PIL import Image
import pytesseract
# import cv2 -> Ya no es necesario para el preprocesamiento simple
# import numpy as np -> Ya no es necesario
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv # Para cargar .env localmente
import google.generativeai as genai # Importar Gemini

# --- Cargar Variables de Entorno y Configurar Gemini ---
load_dotenv() # Carga variables de .env si existe (para desarrollo local)
API_KEY = os.getenv("GOOGLE_API_KEY") # Lee desde variable de entorno

# Configurar Gemini al inicio de la app
if API_KEY:
    try:
        genai.configure(api_key=API_KEY)
        print("API de Google AI configurada correctamente.")
    except Exception as e:
        print(f"Error al configurar la API de Google AI: {e}", file=sys.stderr)
        # La app podría seguir funcionando pero las llamadas a Gemini fallarán.
        # Podrías decidir lanzar un error aquí si Gemini es esencial.
        # raise RuntimeError(f"No se pudo configurar Google AI: {e}") from e
        API_KEY = None # Marcar como no configurada
else:
    print("Advertencia: GOOGLE_API_KEY no encontrada en el entorno.", file=sys.stderr)
    # Considera si la app debe fallar si no hay API Key
    # raise RuntimeError("GOOGLE_API_KEY es requerida pero no se encontró.")

# --- Funciones de Extracción (V6) ---

def clean_ocr_text_for_llm(text: str) -> str:
    """Limpia texto OCR para LLM."""
    if not text: return ""
    text = re.sub(r'\n{2,}', '\n', text)
    text = re.sub(r'[ \t]+', ' ', text)
    text = text.replace('N? Transf.', 'N° Transf.')
    text = text.replace('N* de cuenta', 'N° de cuenta')
    text = text.replace('N* Operación', 'N° Operación')
    text = text.replace('N9 de cuenta', 'N° de cuenta')
    text = text.replace('N9 de usuario', 'N° de usuario')
    return text.strip()

def perform_ocr(image_bytes: bytes, ocr_config: str = '') -> str | None:
    """Realiza OCR desde bytes de imagen."""
    try:
        custom_config = f'-l spa {ocr_config}'.strip()
        # Usamos PIL para abrir desde bytes y convertir a escala de grises
        img = Image.open(io.BytesIO(image_bytes)).convert('L')
        text = pytesseract.image_to_string(img, config=custom_config)
        print(f"--- OCR completado (config: '{custom_config}') ---")
        # Devolver None si Tesseract no encuentra texto (resultado vacío)
        return text if text and text.strip() else None
    except pytesseract.TesseractNotFoundError:
        print("\nError Crítico: 'tesseract' no encontrado.", file=sys.stderr)
        # Este error es a nivel de sistema, lanzamos excepción para que falle el request
        raise RuntimeError("Tesseract OCR no está instalado o no se encuentra en el PATH.")
    except Exception as e:
        print(f"Error durante OCR: {e}", file=sys.stderr)
        return None # Devolver None para errores específicos de imagen

def extract_data_with_gemini(ocr_text: str, filename: str) -> dict:
    """Usa Gemini para extraer datos (Prompt v6)."""
    if not API_KEY:
        print("Error: Intento de llamada a Gemini sin API Key configurada.", file=sys.stderr)
        return {"error": "API Key de Google no configurada en el servidor."}
    if not ocr_text or not ocr_text.strip(): return {"error": "Texto OCR vacío."}

    # Asegurarse de que el modelo existe y está configurado
    try:
        model = genai.GenerativeModel('gemini-1.5-flash-latest')
    except Exception as e:
         print(f"Error al inicializar el modelo Gemini: {e}", file=sys.stderr)
         return {"error": f"Fallo al inicializar modelo Gemini: {e}"}

    cleaned_text = clean_ocr_text_for_llm(ocr_text)

    # ---- Prompt v6 (Copiado directamente) ----
    prompt = f"""
    Analiza el siguiente texto OCR de un comprobante chileno ('{filename}') y extrae la info en JSON. Sigue ESTRICTAMENTE la estructura. Usa el valor JSON `null` si el dato no está CLARO. NO INVENTES.

    **Instrucciones Generales:**
    1.  **Tipo de Operación:** Primero determina si es: A) Transferencia Normal, B) Pago a Comercio, C) Retiro.
    2.  **Roles según Tipo:** Asigna Remitente/Destinatario según el tipo (A, B, C).
    3.  **Banco Origen App:** Busca nombre/logo al inicio/título. Si no, banco remitente o `null`.

    **Instrucciones Campos Específicos:**
    *   **Monto:** IGNORA puntos, comas y ceros finales (,00/.00). Extrae SÓLO el número entero. Ej: "$ 45.890,00" -> 45890.
    *   **RUT vs. Cuenta:** Si formato RUT -> `rut`. Etiquetado "Cuenta N°" -> `cuenta`. Número largo sin formato RUT bajo nombre -> `cuenta`. Número sin guion bajo "Usuario" -> `rut`. Enmascarado -> `rut`. CuentaRUT: rut puede ser cuenta. **NO CONFUNDIR**. Usa `null` si dudas.
    *   **Banco:** Asocia con remitente ("Origen") o destinatario ("Destino").
    *   **Tipo Cuenta:** Busca CuentaRUT, Corriente, Vista, Ahorro. "Producto origen Ahorro" -> `remitente.tipo_cuenta`.
    *   **Fecha/Hora:** DD/MM/YYYY, etc. HH:MM:SS, etc. Convierte "DD Dic YYYY" a "DD/12/YYYY".
    *   **Codigo Transaccion:** Busca "N° Transacción/Operación/Comprobante", etc. También busca números largos (>8 dígitos) aislados arriba (Ej: MP).
    *   **Estado:** "Exitosa", "Realizado", "Aprobada", "Procesado", etc.

    Texto del Comprobante (de {filename}):
    ```
    {cleaned_text}
    ```

    Estructura JSON Requerida (USA JSON `null`, NO `"null"`):
    {{
      "remitente": {{ "nombre": "string | null", "rut": "string | null", "banco": "string | null", "cuenta": "string | null", "tipo_cuenta": "string | null" }},
      "destinatario": {{ "nombre": "string | null", "rut": "string | null", "banco": "string | null", "cuenta": "string | null", "tipo_cuenta": "string | null" }},
      "monto": "number (entero) | null",
      "moneda": "string (CLP) | null",
      "fecha": "string | null",
      "hora": "string | null",
      "asunto": "string | null",
      "codigo_transaccion": "string | null",
      "estado": "string | null",
      "banco_origen_app": "string | null"
    }}

    JSON extraído:
    """
    # ---- Fin Prompt v6 ----

    try:
        generation_config = genai.types.GenerationConfig(temperature=0.05)
        print(f"--- Llamando a la API de Gemini para {filename} ---")
        start_time = time.time()
        response = model.generate_content(prompt, generation_config=generation_config)
        end_time = time.time()
        print(f"--- Llamada completada en {end_time - start_time:.2f} segundos ---")

        # --- Extracción y Parseo (simplificado de V6) ---
        raw_response_text = None
        if hasattr(response, 'parts') and response.parts:
             raw_response_text = "".join(part.text for part in response.parts)
        elif response.candidates and response.candidates[0].finish_reason.name != "STOP":
             reason = response.candidates[0].finish_reason.name
             return {"error": f"Generación detenida por Gemini: {reason}"}
        elif hasattr(response, 'text'): raw_response_text = response.text
        else: return {"error": "Respuesta inválida de Gemini", "raw_response": str(response)}

        if not raw_response_text: return {"error": "Respuesta Gemini sin texto", "raw_response": str(response)}

        json_str = None
        json_match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', raw_response_text, re.DOTALL | re.IGNORECASE)
        if json_match: json_str = json_match.group(1)
        else:
            start = raw_response_text.find('{'); end = raw_response_text.rfind('}')
            if start != -1 and end != -1 and start < end:
                 potential_json = raw_response_text[start:end+1]
                 if ':' in potential_json: json_str = potential_json

        if not json_str: return {"error": "No se extrajo JSON de la respuesta.", "raw_response": raw_response_text}

        try:
            json_str_cleaned = re.sub(r"//.*", "", json_str)
            json_str_cleaned = re.sub(r",\s*([\}\]])", r"\1", json_str_cleaned)
            json_str_cleaned = json_str_cleaned.replace('"null"', 'null')
            extracted_data = json.loads(json_str_cleaned)
            # Post-validación monto
            if 'monto' in extracted_data and isinstance(extracted_data['monto'], (float, str)):
                 try: extracted_data['monto'] = int(float(str(extracted_data['monto']).replace('.', '').replace(',', '')))
                 except (ValueError, TypeError): extracted_data['monto'] = None
            return extracted_data
        except json.JSONDecodeError as json_e:
            return {"error": f"Error parseo JSON: {json_e}", "raw_response": raw_response_text, "json_attempted": json_str_cleaned}

    except Exception as e: # Captura genérica para errores de API
        error_details = str(e)
        print(f"Error llamada API Gemini {filename}: {error_details}", file=sys.stderr)
        # Aquí puedes personalizar qué errores devolver al cliente
        if "API key not valid" in error_details: return {"error": "Error API: Clave inválida."}
        if "429" in error_details or "rate limit" in error_details.lower(): return {"error": "Error API: Límite de tasa excedido."}
        # Devolver un error genérico para otros problemas de API
        return {"error": f"Error inesperado en API Gemini: {error_details}"}


# --- Nueva Función para Aplanar Datos para Excel/Tabla ---
def flatten_data_for_export(gemini_data: dict, filename: str) -> dict:
    """
    Convierte el JSON anidado de Gemini en un diccionario plano
    para usar en Pandas DataFrame y tablas.
    """
    flat_data = {}

    # Manejo de errores en la extracción original
    if "error" in gemini_data:
        return {
            "Archivo": filename,
            "Error Extraccion": gemini_data["error"],
            # Rellenar otros campos con None o algún indicador
            "Fecha y Hora": None, "Origen": None, "Destino": None,
            "Asunto / Descripción": None, "Monto": None, "Estado": None,
            "ID Operacion": None
        }

    # Campos principales
    flat_data["Archivo"] = filename
    flat_data["Fecha y Hora"] = f"{gemini_data.get('fecha', '') or ''} {gemini_data.get('hora', '') or ''}".strip() or None
    flat_data["Asunto / Descripción"] = gemini_data.get('asunto')
    flat_data["Monto"] = gemini_data.get('monto') # Ya debería ser entero
    flat_data["Estado"] = gemini_data.get('estado')
    flat_data["ID Operacion"] = gemini_data.get('codigo_transaccion')
    # Campos derivados/combinados
    remitente = gemini_data.get('remitente', {}) or {}
    destinatario = gemini_data.get('destinatario', {}) or {}
    banco_origen_app = gemini_data.get('banco_origen_app')

    # Lógica para Origen
    origen = remitente.get('nombre') or banco_origen_app or remitente.get('banco') or remitente.get('rut')
    flat_data["Origen"] = origen

    # Lógica para Destino
    destino = destinatario.get('nombre') or destinatario.get('banco') or destinatario.get('rut')
    flat_data["Destino"] = destino

    # Campos detallados (opcional añadirlos aquí o mantenerlos separados)
    # flat_data["Remitente Nombre"] = remitente.get('nombre')
    # flat_data["Remitente RUT"] = remitente.get('rut')
    # flat_data["Remitente Banco"] = remitente.get('banco')
    # flat_data["Destinatario Nombre"] = destinatario.get('nombre')
    # flat_data["Destinatario RUT"] = destinatario.get('rut')
    # flat_data["Destinatario Banco"] = destinatario.get('banco')
    # flat_data["Destinatario Cuenta"] = destinatario.get('cuenta')
    # flat_data["App Origen"] = banco_origen_app

    return flat_data


# --- FastAPI App ---
app = FastAPI(title="Extractor de Comprobantes API")

# Configuración de CORS
origins = ["*"] # Ajusta esto en producción
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- Endpoints Modificados ---

@app.get("/")
async def read_root():
    return {"message": "API Extractor de Comprobantes V6 funcionando!"}

@app.post("/extract")
async def extract_comprobante(file: UploadFile = File(...)):
    """Extrae datos de una única boleta usando OCR y Gemini."""
    if not API_KEY:
        raise HTTPException(status_code=503, detail="Servicio no disponible: API Key no configurada.")

    filename = file.filename or "uploaded_file"
    print(f"Procesando archivo individual: {filename}")

    try:
        contents = await file.read()
        # OCR (usando la función V6)
        ocr_text = perform_ocr(contents) # No necesita config específica aquí

        if ocr_text is None:
             # Intenta de nuevo sin conversión a escala de grises por si acaso
             print("OCR inicial falló, reintentando sin conversión a escala de grises...")
             try:
                 img_color = Image.open(io.BytesIO(contents))
                 ocr_text = pytesseract.image_to_string(img_color, lang='spa')
                 if not ocr_text or not ocr_text.strip(): ocr_text = None # Asegurar que no esté vacío
             except Exception as retry_e:
                 print(f"Reintento de OCR falló: {retry_e}")
                 ocr_text = None

        if ocr_text is None:
            print(f"Error final de OCR para {filename}")
            raise HTTPException(status_code=422, detail="No se pudo extraer texto de la imagen mediante OCR.")

        print(f"Texto OCR obtenido para {filename}:\n---\n{ocr_text[:500]}...\n---")

        # Extracción con Gemini (usando la función V6)
        extracted_data = extract_data_with_gemini(ocr_text, filename)

        # Si Gemini devuelve un error, lo pasamos al cliente
        if "error" in extracted_data:
             # Podríamos mapear errores internos a códigos HTTP apropiados
             status_code = 500 # Error interno genérico
             if "Límite de tasa" in extracted_data["error"]: status_code = 429
             if "API Key" in extracted_data["error"]: status_code = 503
             raise HTTPException(status_code=status_code, detail=extracted_data["error"])

        # Devolver el JSON completo extraído por Gemini
        return JSONResponse(content={
            "filename": filename,
            "extracted_data": extracted_data
            # "ocr_text": ocr_text # Opcional: devolver texto OCR para depuración
        })

    except RuntimeError as rt_e: # Captura errores de Tesseract no encontrado
         raise HTTPException(status_code=500, detail=str(rt_e))
    except Exception as e:
        print(f"Error inesperado procesando {filename}: {e}")
        raise HTTPException(status_code=500, detail=f"Error interno del servidor: {e}")


@app.post("/extract_multi")
async def extract_multiple(files: List[UploadFile] = File(...)):
    """Extrae datos de múltiples boletas (hasta 10) usando OCR y Gemini."""
    if not API_KEY:
        raise HTTPException(status_code=503, detail="Servicio no disponible: API Key no configurada.")
    if len(files) > 10: # Límite
        raise HTTPException(status_code=400, detail="Máximo 10 archivos permitidos.")

    results = []
    for file in files:
        filename = file.filename or f"file_{len(results)+1}"
        print(f"Procesando archivo múltiple: {filename}")
        try:
            contents = await file.read()
            # OCR
            ocr_text = perform_ocr(contents)
            if ocr_text is None:
                # Reintento sin escala de grises (igual que en /extract)
                print(f"OCR inicial falló para {filename}, reintentando...")
                try:
                    img_color = Image.open(io.BytesIO(contents))
                    ocr_text = pytesseract.image_to_string(img_color, lang='spa')
                    if not ocr_text or not ocr_text.strip(): ocr_text = None
                except Exception: ocr_text = None

            if ocr_text:
                # Extracción Gemini
                gemini_result = extract_data_with_gemini(ocr_text, filename)
                results.append({
                    "filename": filename,
                    "extracted_data": gemini_result # Devolvemos el resultado completo de Gemini
                    # "ocr_text": ocr_text # Opcional
                })
            else:
                print(f"Error final de OCR para {filename} en multi-upload.")
                results.append({
                    "filename": filename,
                    "extracted_data": {"error": "OCR falló o no extrajo texto."}
                })
        except RuntimeError as rt_e: # Error Tesseract
             results.append({"filename": filename, "extracted_data": {"error": str(rt_e)}})
             # Podríamos decidir detener todo si Tesseract falla
        except Exception as e:
            print(f"Error inesperado procesando {filename} en multi-upload: {e}")
            results.append({"filename": filename, "extracted_data": {"error": f"Error interno: {e}"}})

    return JSONResponse(content={"results": results})

# Endpoint para exportar datos a Excel (MODIFICADO para usar datos aplanados)
@app.post("/export")
async def export_excel(extracted_results: str = Form(...)): # Recibe el JSON del resultado de /extract_multi
    """
    Exporta los resultados (de /extract o /extract_multi) a Excel.
    Espera un JSON string en 'extracted_results' que contenga una lista de resultados
    o un único resultado como en la salida de los endpoints /extract y /extract_multi.
    """
    try:
        # Parsea el JSON string recibido del formulario
        input_data = json.loads(extracted_results)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=422, detail=f"JSON inválido en extracted_results: {e}")

    processed_rows = []

    # Determina si es un resultado único o múltiple
    if isinstance(input_data, dict) and "extracted_data" in input_data:
        # Asume formato de /extract
        filename = input_data.get("filename", "resultado_unico")
        gemini_data = input_data["extracted_data"]
        processed_rows.append(flatten_data_for_export(gemini_data, filename))
    elif isinstance(input_data, dict) and "results" in input_data:
         # Asume formato de /extract_multi
         for result_item in input_data["results"]:
             filename = result_item.get("filename", "desconocido")
             gemini_data = result_item.get("extracted_data", {"error": "Datos no encontrados"})
             processed_rows.append(flatten_data_for_export(gemini_data, filename))
    elif isinstance(input_data, list):
         # Asume que ya es una lista de resultados de Gemini (menos probable desde Form)
         for i, gemini_data in enumerate(input_data):
              filename = f"item_{i+1}" # Nombre genérico
              processed_rows.append(flatten_data_for_export(gemini_data, filename))
    else:
         raise HTTPException(status_code=422, detail="Formato de 'extracted_results' no reconocido.")


    if not processed_rows:
        raise HTTPException(status_code=400, detail="No hay datos válidos para exportar.")

    # Crear DataFrame con los datos aplanados
    df = pd.DataFrame(processed_rows)

    # Seleccionar y ordenar columnas para la tabla final
    column_order = [
        "Archivo", "Fecha y Hora", "Origen", "Destino",
        "Asunto / Descripción", "Monto", "Estado", "ID Operacion",
        "Error Extraccion" # Incluir columna de error si existe
    ]
    # Filtrar para incluir solo columnas existentes en el DataFrame
    df_export = df[[col for col in column_order if col in df.columns]]


    # Crear Excel en memoria
    stream = io.BytesIO()
    with pd.ExcelWriter(stream, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Extraccion")

        # Aplicar formato de tabla (opcional pero recomendado)
        try:
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo

            worksheet = writer.sheets["Extraccion"]
            num_rows, num_cols = df_export.shape
            if num_rows > 0: # Solo si hay datos
                last_col_letter = get_column_letter(num_cols)
                table_range = f"A1:{last_col_letter}{num_rows + 1}"

                tab = Table(displayName="Extracciones", ref=table_range)
                style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True) # Estilo alternativo
                tab.tableStyleInfo = style
                worksheet.add_table(tab)

                # Autoajustar ancho de columnas (aproximado)
                for col_idx, column in enumerate(df_export.columns):
                    max_length = 0
                    column_letter = get_column_letter(col_idx + 1)
                    # Considera el encabezado y los datos
                    max_length = max(len(str(column)), df_export[column].astype(str).map(len).max())
                    adjusted_width = (max_length + 2) * 1.1 # Añadir padding
                    worksheet.column_dimensions[column_letter].width = adjusted_width

        except ImportError:
            print("Advertencia: openpyxl no encontrado o versión incompatible con tablas. Exportando sin formato de tabla.")
        except Exception as table_e:
            print(f"Advertencia: No se pudo aplicar formato de tabla al Excel: {table_e}")


    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=extraccion_comprobantes.xlsx"}
    )

# Código para ejecutar localmente con Uvicorn (opcional)
# if __name__ == "__main__":
#     import uvicorn
#     uvicorn.run(app, host="0.0.0.0", port=8000)